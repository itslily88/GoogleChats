#!/usr/bin/env python3
# Created by Thomas 'Lily' Lilienthal MAY.2025 - Deschutes County, Oregon - v1
"""
googleChats.py  <parentDirectory>

Walk all sub-folders of <parentDirectory> looking for files named 'messages.json'.
We'll parse those files and create an Excel workbook named googleChats.xlsx
in the directory you passed.

The directory you want to pass should be similar to 
...GoogleChat.Messages_001.001/Google Chat/Groups/
That zip needs to be extracted somewhere first.

requirements:
openpyxl
"""

import sys
import pathlib
from openpyxl import Workbook
import json
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def findMessagesJson(rootDir: pathlib.Path):
    """
    Recursively search rootDir for files called 'messages.json' (case-insensitive).
    Returns a list of pathlib.Path objects.
    """
    return [p for p in rootDir.rglob('*')          # walk every child
            if p.is_file() and p.name.lower() == 'messages.json']

def createWorkbook(outPath: pathlib.Path, jsonFiles, rootDir):
    """
    Make an Excel file with a single sheet and the desired headers.
    Then call parseChats to parse each messages.json in passed directory.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Messages'

    # Attach headers
    headers = ['chatID', 'datetime UTC', 'sender', 'text', 'attachment', 'IP address']
    ws.append(headers)
    
    # Style headers
    headerFont  = Font(bold=True, color='FFFFFF')
    headerFill  = PatternFill(fill_type='solid', fgColor='000000')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = headerFont
        cell.fill = headerFill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Pass worksheet, jsonFiles list, and rootDir and parse
    parseChats(ws, jsonFiles, rootDir)
    cleanup(ws)
    wb.save(outPath)

def parseChats(outWS: Workbook.active, jsonFiles, workbookDir):
    '''
    Loop through jsonFiles and start parsing.
    Data added to worksheet.
    '''
    for jsonPath in jsonFiles:
        # chatID is sourced by the folder it exists in
        chatID = jsonPath.parent.name

        # load the json into variable data
        # Each of these messages.json starts with a key 'messages' that holds everything
        with open(jsonPath, 'r', encoding='utf-8') as f:
            messages = json.load(f).get('messages',[])

        # Loop through each message
        for message in messages:
            sender = message.get('creator',{}).get('email','')
            datetime = message.get('created_date','')

            # Handles if the user updated their message at different times. Finds the actual create
            # MAY.2025, I've only seen this happen on video attachements at this time. Unsure why
            if not datetime: 
                uploads = message.get('previous_message_versions',[{}])
                for upload in uploads:
                    if upload['created_date'] != '': datetime = upload['created_date']

            datetimePretty  = parseGoogleDate(datetime)
            text = message.get('text','')

            # Grabs attachments to handle possible multiple attachments for hyperlink
            attachList = [f.get('export_name', '') for f in message.get('attached_files', [])]
            attachmentNames = '\n'.join(attachList)

            # Grabs an IP address if it exists
            ipAddress = (message.get('upload_metadata', [{}])[0].get('backend_upload_metadata', {}).get('upload_ip')
                         if message.get('upload_metadata') else '')

            # Write found values to worksheet
            outWS.append([chatID, datetimePretty, sender, text, attachmentNames, ipAddress])

            # Makes first attachment in cell a hyperlink relative to where the .xlsx file is
            rowIdx = outWS.max_row
            if attachList:
                firstName = attachList[0]
                absPath   = jsonPath.parent / firstName
                relPath   = os.path.relpath(absPath, start=workbookDir)
                relPath   = relPath.replace('\\', '/')
                cell = outWS.cell(row=rowIdx, column=5)
                cell.value     = attachmentNames
                cell.hyperlink = relPath
                cell.style     = "Hyperlink"

def parseGoogleDate(raw: str):
    '''
    The date held in the json is this weird Friday, October 25, 2024 at 3:20:36 AM UTC
    structure. return will be interpreted as a datetime by excel in UTC
    '''
    # Handle when there is no time stored
    if not raw:
        return ''
    
    # If there is a time, format it
    rawFormat = '%A, %B %d, %Y at %I:%M:%S %p %Z'
    cleaned = raw.replace('\u202f', ' ').strip()
    formatted = datetime.strptime(cleaned, rawFormat)
    return formatted

def cleanup(outWS: Workbook.active):
    '''
    Word Wraps where needed, sets column width, enables filters
    '''
    autofitColumn(outWS, 1) # ChatID
    autofitColumn(outWS, 2) # Datetime
    autofitColumn(outWS, 3) # Email
    outWS.column_dimensions['D'].width = 80 # Text
    outWS.column_dimensions['E'].width = 50 # Attachments
    autofitColumn(outWS, 6) # IP address

    # Prep wordwraps based on maxRow
    wrapAlign = Alignment(wrapText=True)
    maxRow = outWS.max_row

    # set wraps where needed
    outWS[f'D{maxRow}'].alignment = wrapAlign # Text
    outWS[f'E{maxRow}'].alignment = wrapAlign # Attachments

    outWS.auto_filter.ref = f"A1:{get_column_letter(outWS.max_column)}1"
    outWS.freeze_panes = "A2"

def autofitColumn(ws, colIdx, padding=2):
    '''
    Works to autofit columns, where colIdx starts at 1 for col A
    '''
    colLetter = get_column_letter(colIdx)
    maxLen = 0
    for cell in ws[colLetter]:
        if cell.value is None:
            continue
        longest = max(len(str(line)) for line in str(cell.value).split('\n'))
        maxLen = max(maxLen, longest)
    ws.column_dimensions[colLetter].width = maxLen + padding

def main():
    # Grab and validate CLI argument
    if len(sys.argv) != 2:
        print("Usage: python googleChats.py <parentDirectory>")
        sys.exit(1)

    # Establish Root Parent Director to Recursively search through
    rootDir = pathlib.Path(sys.argv[1]).expanduser().resolve()

    # Error if cannot access
    if not rootDir.is_dir():
        print(f"Error: '{rootDir}' is not a directory or cannot be accessed.")
        sys.exit(1)

    # Discover all messages.json files
    jsonFiles = findMessagesJson(rootDir)

    # Close if no messages.json located
    if not jsonFiles:
        print("No messages.json files found under", rootDir)
        sys.exit(0)

    # Report what we found
    print(f"Found {len(jsonFiles)} messages.json file(s). Beginning to parse.")

    # Create basic workbook and start parsing
    outXlsx = rootDir / 'googleChats.xlsx'
    createWorkbook(outXlsx, jsonFiles, rootDir)
    print(f"Created Excel workbook: {outXlsx}")

if __name__ == '__main__':
    main()
